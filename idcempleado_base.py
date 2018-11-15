from openpyxl import load_workbook
import pymysql
import os 
import time
from os import listdir


rootDir = 'xlsx'
lista =[]
Autorizados=[]
count = 0 
lineasf = 2
z = 0 
parar =""
db = "aki"
# f1 = open('./import/lista.csv','w')
# for dirName, subdirList, fileList in os.walk(rootDir):
#     print('Directorio encontrado: %s' % dirName)
#     for fname in fileList:
#         doc = load_workbook(filename = rootDir + '/' + fname)
#         # a = doc.get_sheet_names()
#         hoja = doc['Hoja1']
#         for a in range(2,hoja.max_row):
#         	ccc = hoja['A'+ str (a)].value
#         	Naf = hoja['B'+ str (a)].value
#         	nif = hoja['C'+ str (a)].value
#         	FechaAlta = str(hoja['D'+ str (a)].value)
#         	Autorizado = hoja['E'+ str (a)].value
#         	f1.write(str(str(ccc) + ";" + str(Naf) + ";" + str(nif) + ";" + str(FechaAlta) + ";"+ str(Autorizado) + "\n"))
# f1.close()
# # nos conectamos a la base de dato
# connection = pymysql.connect(host='localhost',port=3306,user='root',db = db)
# cursor = connection.cursor()
# sql = "LOAD DATA INFILE" + "'" + "C:\\\\Users\\\\jflores\\\\Desktop\\\\Programacion\\\\import\\\\" + "lista.csv" + "'" + "INTO TABLE Subvenciones1 CHARACTER SET UTF8 FIELDS TERMINATED BY ';' LINES TERMINATED BY '\n' (CCC, NAF, NIF, FECHAALTA, AUTORIZADO);"
# cursor.execute(sql)
# connection.commit()
# connection.close()
#consultamos las empresas con sus diferentes Autorizados
connection1 = pymysql.connect(host='localhost',port=3306,user='root',db = db)
cursor = connection1.cursor()
sql = "SELECT DISTINCT AUTORIZADO FROM Subvenciones1"
cursor.execute(sql)
filas =cursor.fetchall()
for fila in filas:
    Autorizado= fila[0]
    Autorizados.append(Autorizado)
connection1.close()



connection = pymysql.connect(host='localhost',port=3306,user='root',db = db)
cursor = connection.cursor()
for a in range(0,len(Autorizados)):
    # print(Autorizados[a])
    sql = "SELECT * from subvenciones1 where AUTORIZADO = '" + Autorizados[a]  +  "'"
    nombre = time.strftime("%d%H%M%S")
    id_winsuite = "99R99999"
    autorizacion = Autorizados[a]
    autorizacion = autorizacion.replace("\r","")
    autorizacion = autorizacion.zfill(8)
    f = open('./afi/'+ nombre +'.afi','w')
    f.write('ETIAFI80WS820'+ autorizacion + id_winsuite + "201504221839" + str(nombre) + "AFIN 00000000000000A "+"\n")
    cursor.execute(sql)
    filas1 =cursor.fetchall()
    for fila in filas1:
        CCC = fila[1]
        ccc = "0111" + str(CCC[0:2]) + str(CCC[3:12])
        NAF = fila [2]
        naf =NAF[0:2] + NAF[3:14]
        NIF = fila[3]
        FECHAALTA = fila[4]
        
        ########## comenzamos a crear el AFI ##################################
        j = (NIF[1:2])
        k = NIF[9:10]
        j = j.isdigit()
        if j is True:
            tipo_identificador ="1"
            fl= 1
        k = k.isdigit()
        if k is True:
            tipo_identificador ="2"
            fl = 1
        if fl == 0:
            tipo_identificador = "6"
        FECHAALTA = str(FECHAALTA)
        anho_desde = str((FECHAALTA[0:4]))
        mes_desde = str((FECHAALTA[5:7]))
        dia_desde = str((FECHAALTA[8:10]))
        periodo_desde = str(anho_desde)+str(mes_desde)+str(dia_desde)
        f.write('EMP'+ ccc + "                    " + "000000000000000" + "                 "   +"\n")
        f.write("RZS02"+"                                                       "+"00000000  " +"\n")
        f.write("TRA"+naf+ tipo_identificador + "   "+"0000"+NIF+"                                     "+"\n")
        f.write ("FABIDC00000000000000"+" " + "000" + "  " + "000000" + " " + "0000000000000000" + " "+ "0000" + "   " + "N" + " "+ periodo_desde +"   "+ "\n")

        count = count + 1
        lineasf = int(lineasf) + 4
        countfinal = str(count)
        countfinal = countfinal.zfill(5)
        lineasf = str(lineasf)
        lineasf = lineasf.zfill(8)
        if count == 4000:
            lineasf = str(lineasf)
            f.write( "ETFAFI80WS820" + autorizacion + id_winsuite + "201504221840"  + nombre + "AFIN " +countfinal + lineasf + "   "+"\n")
            z = z + 1
            lineasf = 2
            count = 0
            nombre = int(nombre)
            nombre = nombre + 1
            nombre = str(nombre)
            f = open('./afi/'+ nombre +'.afi','w')
            f.write('ETIAFI80WS820'+ autorizacion + id_winsuite + "201504221839" + nombre + "AFIN 00000000000000A "+"\n")
        lineasf = str(lineasf)
    
    f.write( "ETFAFI80WS820" + autorizacion + id_winsuite + "201504221840"  + str(nombre) + "AFIN " +str(countfinal) + str(lineasf) + "   "+"\n")
    count = 0 
    lineasf = 0 
    time.sleep(1)